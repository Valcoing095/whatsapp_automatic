import os
import time
import urllib.parse

from flask import Flask, jsonify, render_template, request, session
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

app = Flask(__name__)
app.secret_key = os.urandom(32)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

driver = None


def process_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    header_lower = [str(h).lower() if h else "" for h in headers]

    required = ["nombre", "celular", "placa"]
    if not all(any(req in h for h in header_lower) for req in required):
        return None, "El Excel debe tener columnas: nombre, celular, placa"

    idx_nombre = next(i for i, h in enumerate(header_lower) if "nombre" in h)
    idx_celular = next(i for i, h in enumerate(header_lower) if "celular" in h)
    idx_placa = next(i for i, h in enumerate(header_lower) if "placa" in h)

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_nombre]:
            celular = str(int(row[idx_celular])) if row[idx_celular] else ""
            placa = str(row[idx_placa]) if len(row) > idx_placa and row[idx_placa] else ""
            data.append({"nombre": str(row[idx_nombre]), "celular": celular, "placa": placa})

    return data, None


def format_phone(phone):
    phone = "".join(c for c in str(phone).strip() if c.isdigit())
    if not phone.startswith("57") and len(phone) == 10:
        phone = "57" + phone
    elif not phone.startswith("57") and len(phone) == 9:
        phone = "5" + phone
    return phone


def get_chrome_driver():
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("detach", True)
    return webdriver.Chrome(options=options)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No se encontró archivo"}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Archivo inválido"}), 400

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    data, error = process_excel(filepath)
    if error:
        return jsonify({"error": error}), 400

    session["clientes"] = data
    return jsonify({"clientes": data, "total": len(data)})


@app.route("/iniciar-whatsapp", methods=["POST"])
def iniciar_whatsapp():
    global driver
    if not driver:
        driver = get_chrome_driver()
        driver.get("https://web.whatsapp.com")
    return jsonify({"status": "WhatsApp Web iniciado. Escanea el QR code."})


@app.route("/enviar", methods=["POST"])
def enviar():
    global driver

    data = request.json
    clientes = data.get("clientes", [])
    mensaje_template = data.get("mensaje", "")

    if not driver:
        return jsonify({"success": False, "error": "WhatsApp no está iniciado"}), 400

    try:
        driver.current_url
    except:
        driver = get_chrome_driver()
        driver.get("https://web.whatsapp.com")
        return jsonify({"success": False, "error": "WhatsApp reiniciado. Escanea QR nuevamente."}), 400

    resultados = []
    for cliente in clientes:
        nombre = cliente.get("nombre", "")
        celular = cliente.get("celular", "")
        placa = cliente.get("placa", "")

        mensaje = mensaje_template.replace("{nombre}", nombre).replace("{placa}", placa).replace("{celular}", celular)
        phone = format_phone(celular)

        encoded_msg = urllib.parse.quote(mensaje)
        driver.get(f"https://web.whatsapp.com/send?phone={phone}&text={encoded_msg}")
        time.sleep(4)

        send_btn = None
        for xpath in [
            "//span[@data-icon='send']",
            "//button[@data-tab='10']",
            "//span[contains(@data-icon, 'send')]",
            "//div[@role='button' and contains(@aria-label, 'Enviar')]",
            "//button[contains(@aria-label, 'send')]",
        ]:
            try:
                send_btn = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, xpath))
                )
                if send_btn:
                    break
            except:
                continue

        status = "enviado" if send_btn else "fallido"
        if send_btn:
            send_btn.click()
            time.sleep(1.5)

        resultados.append({"nombre": nombre, "celular": celular, "placa": placa, "status": status})

    enviados = sum(1 for r in resultados if r["status"] == "enviado")
    return jsonify({"success": True, "resultados": resultados, "enviados": enviados, "total": len(clientes)})


@app.route("/cerrar", methods=["POST"])
def cerrar():
    global driver
    if driver:
        driver.quit()
        driver = None
    return jsonify({"status": "cerrado"})


if __name__ == "__main__":
    app.run(debug=True, port=5000)